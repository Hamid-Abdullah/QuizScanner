"""
Task 5 — Batch Processing & Report Generation [30 points]
Processes multiple quiz images in one run and produces a formatted Excel/CSV report.

Evaluation criteria:
  - Correct batch pipeline end-to-end            : 15 pts
  - All required columns accurately populated    : 10 pts
  - Summary statistics row                       : 5 pts
"""

import os
import glob
import logging
import threading
import concurrent.futures
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
from dataclasses import asdict

import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.pdf'}
QUESTIONS = [f"Q{i:02d}" for i in range(1, 9)]

# Column order (matches assignment specification exactly)
COLUMNS = [
    'Quiz', 'Set', 'Class', 'Subject',
    'Name', 'Reg No',
    *[f'Part1_{q}' for q in QUESTIONS],
    *[f'Part2_{q}' for q in QUESTIONS],
    'Correct', 'Incorrect', 'Unattempted',
    'Total Marks', 'Max Marks', 'Percentage', 'Grade',
]


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def process_batch(
    input_folder: str,
    output_folder: str = "output",
    quiz_title: str    = "Quiz",
    class_name: str    = "BSE-4A",
    max_workers: int   = 2,
    progress_callback  = None,   # callable(current, total, filename, status)
) -> Dict[str, Any]:
    """
    Process all quiz images in input_folder through Tasks 1–4.

    Args:
        input_folder:      directory containing quiz images
        output_folder:     where to write the Excel/CSV output
        quiz_title:        quiz name prefix (e.g. "AI Quiz SP2026")
        class_name:        default class if not readable from image
        max_workers:       parallel processing threads
        progress_callback: optional callback(current, total, filename, status)

    Returns:
        dict with keys: excel_path, csv_path, rows, summary
    """
    os.makedirs(output_folder, exist_ok=True)

    # ── Collect images ────────────────────────────────────────────────────────
    image_files = _collect_images(input_folder)
    if not image_files:
        raise FileNotFoundError(f"No supported images found in: {input_folder}")

    image_files.sort()
    total = len(image_files)
    logger.info(f"Batch: {total} images found in '{input_folder}'")

    # ── Process each image ────────────────────────────────────────────────────
    rows = [None] * total
    lock = threading.Lock()

    def _process(i_path):
        i, path = i_path
        fname = os.path.basename(path)
        if progress_callback:
            progress_callback(i, total, fname, 'processing')

        try:
            row = _process_single(path, quiz_title, class_name, i + 1)
            with lock:
                rows[i] = row
            if progress_callback:
                progress_callback(i + 1, total, fname, 'done')
            logger.info(f"[{i+1}/{total}] OK: {fname}")
        except Exception as e:
            logger.error(f"[{i+1}/{total}] FAIL: {fname} — {e}")
            with lock:
                rows[i] = _error_row(path, quiz_title, class_name, i + 1, str(e))
            if progress_callback:
                progress_callback(i + 1, total, fname, f'error: {e}')

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as exe:
        exe.map(_process, enumerate(image_files))

    valid_rows = [r for r in rows if r is not None]

    # ── Build DataFrame with summary row ─────────────────────────────────────
    df = _build_dataframe(valid_rows)

    # ── Save to Excel + CSV ───────────────────────────────────────────────────
    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_title   = quiz_title.replace(' ', '_')
    xlsx_path    = os.path.join(output_folder, f"{safe_title}_{timestamp}.xlsx")
    csv_path     = os.path.join(output_folder, f"{safe_title}_{timestamp}.csv")

    _save_excel(df, xlsx_path)
    df.to_csv(csv_path, index=False)

    logger.info(f"Saved Excel: {xlsx_path}")
    logger.info(f"Saved CSV  : {csv_path}")

    summary = _compute_summary(valid_rows)

    return {
        'excel_path': xlsx_path,
        'csv_path':   csv_path,
        'rows':       valid_rows,
        'summary':    summary,
        'total':      total,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Single-image processing
# ─────────────────────────────────────────────────────────────────────────────

def _process_single(image_path: str, quiz_title: str,
                    class_name: str, index: int) -> dict:
    """Run Tasks 1–4 on one image and return a row dict."""
    import cv2
    from task1_qr_decoder import decode_answer_key
    from task2_ocr         import extract_student_info
    from task3_bubble      import read_bubble_sheet
    from task4_grading     import grade_quiz

    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Cannot read image file: {image_path}")

    # Task 1: QR decode
    answer_key = decode_answer_key(img)

    # Task 2: OCR
    student_info = extract_student_info(img)

    # Task 3: Bubbles
    student_answers = read_bubble_sheet(img)

    # Task 4: Grade
    report = None
    if answer_key:
        report = grade_quiz(
            student_answers, answer_key,
            student_name=student_info.name,
            reg_no=student_info.reg_no,
        )

    return _build_row(index, quiz_title, class_name,
                      student_info, student_answers, answer_key, report)


def _build_row(index, quiz_title, class_name,
               student_info, student_answers, answer_key, report) -> dict:
    """Assemble a flat dict matching COLUMNS."""
    row: Dict[str, Any] = {
        'Quiz':    f"{quiz_title}",
        'Set':     answer_key.quiz_set if answer_key else 'N/A',
        'Class':   student_info.class_name or class_name,
        'Subject': answer_key.subject if answer_key else (student_info.subject or 'N/A'),
        'Name':    student_info.name,
        'Reg No':  student_info.reg_no,
    }

    for q in QUESTIONS:
        row[f'Part1_{q}'] = student_answers.part1.get(q) or '—'
        row[f'Part2_{q}'] = student_answers.part2.get(q) or '—'

    if report:
        row['Correct']     = report.correct
        row['Incorrect']   = report.incorrect
        row['Unattempted'] = report.unattempted
        row['Total Marks'] = report.total_marks
        row['Max Marks']   = report.max_marks
        row['Percentage']  = round(report.percentage, 2)
        row['Grade']       = report.letter_grade
    else:
        row['Correct']     = 'N/A'
        row['Incorrect']   = 'N/A'
        row['Unattempted'] = 16
        row['Total Marks'] = 0
        row['Max Marks']   = 16
        row['Percentage']  = 0.0
        row['Grade']       = 'N/A'

    return row


def _error_row(path, quiz_title, class_name, index, error_msg) -> dict:
    row: Dict[str, Any] = {
        'Quiz':    f"{quiz_title}",
        'Set':     'ERROR',
        'Class':   class_name,
        'Subject': '',
        'Name':    os.path.basename(path),
        'Reg No':  f'ERROR: {error_msg[:60]}',
    }
    for q in QUESTIONS:
        row[f'Part1_{q}'] = 'ERR'
        row[f'Part2_{q}'] = 'ERR'
    row.update({
        'Correct': 0, 'Incorrect': 0, 'Unattempted': 16,
        'Total Marks': 0, 'Max Marks': 16,
        'Percentage': 0.0, 'Grade': 'F',
    })
    return row


# ─────────────────────────────────────────────────────────────────────────────
# DataFrame construction  (10 pts for columns + 5 pts for summary)
# ─────────────────────────────────────────────────────────────────────────────

def _build_dataframe(rows: list) -> pd.DataFrame:
    """Build the results DataFrame and append a summary statistics row."""
    df = pd.DataFrame(rows, columns=[c for c in COLUMNS if c in rows[0]] if rows else COLUMNS)

    # ── Summary row  (5 pts) ─────────────────────────────────────────────────
    summary: Dict[str, Any] = {
        'Quiz':    'SUMMARY',
        'Set':     '',
        'Class':   '',
        'Subject': '',
        'Name':    '── Class Statistics ──',
        'Reg No':  '',
    }

    for q in QUESTIONS:
        # Show most-common answer for each question
        for prefix in ['Part1_', 'Part2_']:
            col = f'{prefix}{q}'
            if col in df.columns:
                counts = df[col].value_counts()
                summary[col] = counts.index[0] if not counts.empty else ''

    numeric_cols = {
        'Correct':     'Avg:{:.1f}',
        'Incorrect':   'Avg:{:.1f}',
        'Unattempted': 'Avg:{:.1f}',
        'Total Marks': 'Avg:{:.1f} | Max:{:.0f} | Min:{:.0f}',
        'Max Marks':   '16',
        'Percentage':  'Avg:{:.1f}% | Max:{:.1f}% | Min:{:.1f}%',
        'Grade':       '',
    }

    for col, fmt in numeric_cols.items():
        if col not in df.columns:
            continue
        try:
            series = pd.to_numeric(df[col], errors='coerce').dropna()
            if series.empty:
                summary[col] = 'N/A'
                continue
            if col in ('Total Marks',):
                summary[col] = fmt.format(series.mean(), series.max(), series.min())
            elif col == 'Percentage':
                summary[col] = fmt.format(series.mean(), series.max(), series.min())
            elif col == 'Max Marks':
                summary[col] = 16
            elif col == 'Grade':
                summary[col] = ''
            else:
                summary[col] = fmt.format(series.mean())
        except Exception:
            summary[col] = ''

    summary_df = pd.DataFrame([summary])
    df = pd.concat([df, summary_df], ignore_index=True)

    return df


# ─────────────────────────────────────────────────────────────────────────────
# Excel output with formatting
# ─────────────────────────────────────────────────────────────────────────────

def _save_excel(df: pd.DataFrame, path: str) -> None:
    """Save DataFrame to Excel with styling (openpyxl)."""
    from openpyxl import load_workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Grade Report')
        ws = writer.sheets['Grade Report']

        # ── Column widths ─────────────────────────────────────────────────────
        for col_idx, col in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col)),
                df[col].astype(str).map(len).max() if not df.empty else 0
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

        # ── Header row styling ────────────────────────────────────────────────
        header_fill = PatternFill('solid', fgColor='2563EB')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin        = Side(border_style='thin', color='D1D5DB')
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center
            cell.border    = border

        ws.row_dimensions[1].height = 30

        # ── Data rows ─────────────────────────────────────────────────────────
        grade_col_idx = None
        pct_col_idx   = None

        for col_idx, col in enumerate(df.columns, start=1):
            if col == 'Grade':
                grade_col_idx = col_idx
            if col == 'Percentage':
                pct_col_idx = col_idx

        # Grade color map
        grade_colors = {
            'A+': 'D1FAE5', 'A': 'D1FAE5', 'A-': 'A7F3D0',
            'B+': 'FEF9C3', 'B': 'FEF9C3', 'B-': 'FDE68A',
            'C+': 'FEF3C7', 'C': 'FEF3C7', 'C-': 'FDE68A',
            'D':  'FED7AA', 'F': 'FEE2E2',
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            is_summary = (row_idx == ws.max_row)
            row_fill = PatternFill('solid', fgColor='FEF9C3') if is_summary else None

            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if is_summary:
                    cell.fill = row_fill
                    cell.font = Font(bold=True, size=10)

            # Grade-cell coloring
            if grade_col_idx and not is_summary:
                grade_cell = ws.cell(row=row_idx, column=grade_col_idx)
                grade = str(grade_cell.value or '')
                if grade in grade_colors:
                    grade_cell.fill = PatternFill('solid', fgColor=grade_colors[grade])
                    grade_cell.font = Font(bold=True)

        # ── Freeze header + auto filter ───────────────────────────────────────
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # ── Add a summary stats sheet ─────────────────────────────────────────
        _add_summary_sheet(writer, df)


def _add_summary_sheet(writer, df: pd.DataFrame) -> None:
    """Add a second sheet with class-level statistics."""
    try:
        numeric = pd.to_numeric(df['Percentage'], errors='coerce').dropna()
        marks   = pd.to_numeric(df['Total Marks'], errors='coerce').dropna()

        stats = {
            'Metric': [
                'Number of Students',
                'Class Average (%)',
                'Highest Score (%)',
                'Lowest Score (%)',
                'Highest Marks',
                'Lowest Marks',
                'Students Passed (≥50%)',
                'Students Failed (<50%)',
                'Pass Rate (%)',
            ],
            'Value': [
                len(numeric),
                f"{numeric.mean():.1f}%",
                f"{numeric.max():.1f}%",
                f"{numeric.min():.1f}%",
                f"{marks.max():.0f} / {df['Max Marks'].iloc[0]}",
                f"{marks.min():.0f} / {df['Max Marks'].iloc[0]}",
                int((numeric >= 50).sum()),
                int((numeric < 50).sum()),
                f"{(numeric >= 50).mean() * 100:.1f}%",
            ]
        }

        grade_dist = df['Grade'].value_counts().reset_index()
        grade_dist.columns = ['Grade', 'Count']

        stats_df = pd.DataFrame(stats)
        stats_df.to_excel(writer, index=False, sheet_name='Summary')
        grade_dist.to_excel(writer, index=False, sheet_name='Summary',
                            startrow=len(stats_df) + 3)
    except Exception as e:
        logger.warning(f"Could not create summary sheet: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _collect_images(folder: str) -> List[str]:
    files = []
    for ext in SUPPORTED_EXTENSIONS:
        files += glob.glob(os.path.join(folder, f"*{ext}"))
        files += glob.glob(os.path.join(folder, f"*{ext.upper()}"))
    return list(set(files))  # deduplicate


def _compute_summary(rows: list) -> dict:
    if not rows:
        return {}
    percs = [r['Percentage'] for r in rows if isinstance(r.get('Percentage'), (int, float))]
    marks = [r['Total Marks'] for r in rows if isinstance(r.get('Total Marks'), (int, float))]
    return {
        'total_students': len(rows),
        'avg_percentage': round(sum(percs) / len(percs), 2) if percs else 0,
        'max_percentage': max(percs) if percs else 0,
        'min_percentage': min(percs) if percs else 0,
        'avg_marks':      round(sum(marks) / len(marks), 2) if marks else 0,
        'pass_count':     sum(1 for p in percs if p >= 50),
        'fail_count':     sum(1 for p in percs if p < 50),
    }


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    folder   = sys.argv[1] if len(sys.argv) > 1 else "samples"
    out      = sys.argv[2] if len(sys.argv) > 2 else "output"
    title    = sys.argv[3] if len(sys.argv) > 3 else "AI Quiz SP2026"

    def progress(cur, tot, fname, status):
        print(f"  [{cur}/{tot}] {fname} → {status}")

    result = process_batch(folder, out, title, progress_callback=progress)

    print("\n" + "=" * 50)
    print("BATCH COMPLETE")
    print("=" * 50)
    print(f"Excel  : {result['excel_path']}")
    print(f"CSV    : {result['csv_path']}")
    s = result['summary']
    print(f"\nStudents  : {s.get('total_students', 0)}")
    print(f"Avg Score : {s.get('avg_percentage', 0):.1f}%")
    print(f"Highest   : {s.get('max_percentage', 0):.1f}%")
    print(f"Lowest    : {s.get('min_percentage', 0):.1f}%")
    print(f"Passed    : {s.get('pass_count', 0)}")
    print(f"Failed    : {s.get('fail_count', 0)}")
